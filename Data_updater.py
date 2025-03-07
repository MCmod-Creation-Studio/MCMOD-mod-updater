# 导入所需的模块和库
import config
from datetime import datetime
import PyTaskbar
import requests as rq
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from rich.progress import Progress
from urllib3 import disable_warnings as urllib3_disable_warnings
import concurrent.futures
# 导入自定义下载器模块
from Mod_downloader import download_mod
import time

config = config.Config()
DATABASE_PATH = config.DATABASE_PATH
download_enable = config.download_enable
POOL_SIZE = config.POOL_SIZE

session = rq.Session()
adapter = rq.adapters.HTTPAdapter(pool_connections=POOL_SIZE, pool_maxsize=POOL_SIZE, max_retries=3)
session.mount('https://', adapter)
session.mount('http://', adapter)

headers = config.headers

# 忽略urllib3的警告
urllib3_disable_warnings()
# 定义填充颜色
green_fill = PatternFill(patternType='solid', fgColor='90EE90')  # 淡绿色
red_fill = PatternFill(patternType='solid', fgColor='D94600')  # 红色

# 初始化一些全局变量
max_rowFIX = unmatchedSum = matchedSum = rq.session().keep_alive = False
DuplicatesList = []
processesTime = datetime.now()
validify_processesTime = str(processesTime).replace(" ", "+").replace(":", "-")
processes_download_mark = False

# 检查环境变量是否设置了DATABASE_PATH
if not DATABASE_PATH:
    raise Exception('DATABASE_PATH environment variable not set.')

# 加载Excel工作簿
wb = load_workbook(DATABASE_PATH)
exl = wb.active

# 初始化任务栏进度条
prog = PyTaskbar.Progress()
prog.init()
prog.setState('loading')


# 通过列和行获取Excel单元格的值
def Vexl(column, row):
    return exl[f"{column}{row}"].value


# 更新时间信息
def time_update(type: str, row, time):
    Row_Str = str(row)
    if type == "Latest":
        exl["H" + Row_Str] = Vexl("G", row)
        exl["G" + Row_Str] = Vexl("F", row)
        exl["F" + Row_Str] = str(time)
    elif type == "Json":
        exl["L" + Row_Str] = Vexl("K", row)
        exl["K" + Row_Str] = Vexl("J", row)
        exl["J" + Row_Str] = str(time)


# 检查是否存在重复项
def check_duplicates(num_id):
    if Vexl("F", num_id) != Vexl("G", num_id):
        exl[f"F{num_id}"].fill = red_fill
        return False
    else:
        exl[f"F{num_id}"].fill = green_fill
        return True


# 遍历Excel表格，获取最大行数
for max_rowFIX, row in enumerate(exl, 1):
    if all(c.value is None for c in row):
        break
max_rowFIX -= 2


# 获取Curseforge项目的API JSON数据
def get_cfwidget_api_json(cf_project_id):
    retry_count = 0
    max_retries = config.TIMEOUT_RETRY

    while retry_count < max_retries:
        try:
            return session.get(f'https://api.cfwidget.com/{cf_project_id}',
                               headers=headers, params={'param': '1'}, verify=False).json()
        except Exception as e:
            retry_count += 1
            if retry_count >= max_retries:
                print(f"Error getting Curseforge data for {cf_project_id} after {max_retries} attempts: {e}")
                return {"files": []}
            else:
                wait_time = 2 ** retry_count  # Exponential backoff
                print(f"Attempt {retry_count} failed for {cf_project_id}, retrying in {wait_time} seconds...")
                time.sleep(wait_time)


def get_modrinth_api_json(mr_project_id):
    retry_count = 0
    max_retries = config.TIMEOUT_RETRY

    while retry_count < max_retries:
        try:
            return session.get(f'https://api.modrinth.com/v2/project/{mr_project_id}',
                               params={'param': '1'}, headers=headers, verify=False).json()
        except Exception as e:
            retry_count += 1
            if retry_count >= max_retries:
                print(f"Error getting Modrinth data for {mr_project_id} after {max_retries} attempts: {e}")
                return {"versions": [], "updated": "Error"}
            else:
                wait_time = 2 ** retry_count  # Exponential backoff
                print(f"Attempt {retry_count} failed for {mr_project_id}, retrying in {wait_time} seconds...")
                time.sleep(wait_time)


def process_mod(num_id):
    mod_name = Vexl("C", num_id)
    latest_time = ""
    file_ids = ""

    try:
        curseforge_id = Vexl("D", num_id)
        modrinth_id = Vexl("E", num_id)

        if curseforge_id is not None:
            website = "Curseforge"
            if not str(curseforge_id).count("/"):
                json_data = get_cfwidget_api_json(curseforge_id)['files']
                file_ids = [i["id"] for i in json_data]
                latest_time = str(dict(json_data[0])['uploaded_at'])
            else:
                latest_time = ""
                for m in str(curseforge_id).split("/"):
                    json_data = get_cfwidget_api_json(m)['files']
                    file_ids = [i["id"] for i in json_data]
                    latest_time += str(dict(json_data[0])['uploaded_at']) + "  "

        elif modrinth_id is not None:
            website = "Modrinth"
            if not str(modrinth_id).count("/"):
                json_data = get_modrinth_api_json(modrinth_id)
                file_ids = json_data["versions"]
                latest_time = str(json_data['updated'])
            else:
                latest_time = ""
                file_ids = []
                for m in str(modrinth_id).split("/"):
                    json_data = get_modrinth_api_json(m)
                    file_ids += json_data["versions"]
                    latest_time += str(json_data['updated']) + "  "
        else:
            latest_time = "读取失败"
            website = None

        # Update time and check for duplicates
        time_update("Latest", num_id, latest_time)
        time_update("Json", num_id, file_ids)
        matched = check_duplicates(num_id)

        return {
            "num_id": num_id,
            "mod_name": mod_name,
            "latest_time": latest_time,
            "file_ids": file_ids,
            "website": website if 'website' in locals() else None,
            "id": curseforge_id or modrinth_id,
            "matched": matched
        }

    except Exception as err:
        print(f"{num_id - 1} 出现错误:{err}，跳过该项")
        return {
            "num_id": num_id,
            "mod_name": mod_name,
            "latest_time": "读取过程出现错误",
            "file_ids": "",
            "matched": False
        }


# 处理最新上传的模组信息
def latest_upload():
    global DuplicatesList, matchedSum, unmatchedSum, processes_download_mark

    with Progress() as progress:
        task = progress.add_task("[red]正在初始化......", total=max_rowFIX)

        # Use ThreadPoolExecutor with pool size from config
        with concurrent.futures.ThreadPoolExecutor(max_workers=POOL_SIZE) as executor:
            # Submit all tasks to the executor
            future_to_mod = {executor.submit(process_mod, num_id): num_id
                             for num_id in range(2, max_rowFIX + 2)}

            # Process results as they complete
            for future in concurrent.futures.as_completed(future_to_mod):
                num_id = future_to_mod[future]
                progress.update(task, advance=1,
                                description=f"[yellow]f[ {num_id - 1}/{max_rowFIX} ]处理中...")
                prog.setProgress(int(round(num_id * 100 / max_rowFIX, 0)))

                try:
                    result = future.result()
                    if not result["matched"]:
                        DuplicatesList.append(
                            f"{num_id}: {result['mod_name']} || {result['latest_time']} | {Vexl('G', num_id)}")
                        unmatchedSum += 1

                        # Download if needed
                        if download_enable and 'website' in result and result['website']:
                            # 寻找最近得有效list作为pastJson
                            if not (eval(Vexl("K", num_id)) == list):
                                pastJson = Vexl("K", num_id)
                            elif not (eval(Vexl("L", num_id)) == list):
                                pastJson = Vexl("L", num_id)
                            elif not (eval(Vexl("J", num_id)) == list):
                                pastJson = Vexl("M", num_id)
                            else:
                                pastJson = None

                            if pastJson:
                                fileID = list(set(result['file_ids']).difference(set(pastJson)))
                                if fileID:
                                    download_mod(result['website'], Vexl("B", num_id),
                                                 validify_processesTime, result['id'], fileID)
                                    processes_download_mark = True
                    else:
                        matchedSum += 1

                    print(num_id - 1, result['mod_name'], result['latest_time'])

                except Exception as exc:
                    print(f'{num_id - 1} generated an exception: {exc}')

        progress.update(task, advance=0, description="[green][ √ ]已完成处理！")


# 初始化时更新时间
time_update("Latest", 1, processesTime)
time_update("Json", 1, processesTime)
# 开始处理最新上传的模组信息
latest_upload()

# 打印总结信息
print(f"匹配：{matchedSum}，不匹配{unmatchedSum}")
print(datetime.now(), "完成遍历")
if DuplicatesList:
    print(f"发现已更新的模组！：\n {DuplicatesList}")
else:
    print("没有发现已更新的模组")
# 保存Excel文件
try:
    wb.save(DATABASE_PATH)
    print("所有改动已成功保存")
except Exception as E:
    new_path = DATABASE_PATH.replace(".xlsx", f"_{datetime.now().strftime('%Y%m%d%H%M%S')}.xlsx")
    print("保存出错: {0}\n 将保存至新文件{1}.xlsx".format(E, new_path))
    wb.save(new_path)
# 更新任务栏状态
prog.setState('done')
