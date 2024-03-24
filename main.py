# 导入所需的模块和库
from os import getenv
from datetime import datetime
import PyTaskbar
import requests as rq
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from rich.progress import Progress
from urllib3 import disable_warnings as urllib3_disable_warnings

# 导入自定义下载器模块
from Mod_downloader import download_mod

# 忽略urllib3的警告
urllib3_disable_warnings()
# 定义填充颜色
green_fill = PatternFill(patternType='solid', fgColor='90EE90')  # 淡绿色
red_fill = PatternFill(patternType='solid', fgColor='D94600')  # 红色

# 初始化一些全局变量
max_rowFIX = unmatchedSum = matchedSum = rq.session().keep_alive = False
DuplicatesList = []
processesTime = datetime.now()
# 设置请求头
headers = {
    'Referer': '',
    'If-None-Match': '"LMyW4mgAfp2S0ragPuZKjIpMkas="',
    'If-Modified-Since': 'Wed, 25 Oct 2020 09:37:56 GMT',
    'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/119.0.0.0 Safari/537.36 Edg/119.0.0.0',
}

# 检查环境变量是否设置了DATABASE_PATH
if not getenv('DATABASE_PATH'):
    raise Exception('DATABASE_PATH environment variable not set.')
DatabasePath = getenv('DATABASE_PATH')
# 加载Excel工作簿
wb = load_workbook(DatabasePath)
exl = wb.active

# 初始化任务栏进度条
prog = PyTaskbar.Progress()
prog.init()
prog.setState('loading')


# 通过列和行获取Excel单元格的值
def Vexl(column, row):
    return exl[f"{column}{row}"].value


# 更新时间信息
def time_update(type: str, row, time):
    Row_Str = str(row)
    if type == "Latest":
        exl["H" + Row_Str] = Vexl("G", row)
        exl["G" + Row_Str] = Vexl("F", row)
        exl["F" + Row_Str] = str(time)
    elif type == "Json":
        exl["L" + Row_Str] = Vexl("K", row)
        exl["K" + Row_Str] = Vexl("J", row)
        exl["J" + Row_Str] = str(time)


# 检查是否存在重复项
def check_duplicates(num_id):
    if Vexl("F", num_id) != Vexl("G", num_id):
        exl[f"F{num_id}"].fill = red_fill
        return False
    else:
        exl[f"F{num_id}"].fill = green_fill
        return True


# 遍历Excel表格，获取最大行数
for max_rowFIX, row in enumerate(exl, 1):
    if all(c.value is None for c in row):
        break
max_rowFIX -= 2


# 获取Curseforge项目的API JSON数据
def get_cfwidget_api_json(cf_project_id):
    global headers
    return rq.get(f'https://api.cfwidget.com/{cf_project_id}', headers=headers, params={'param': '1'},
                  verify=False).json()


# 获取Modrinth项目的API JSON数据
def get_modrinth_api_json(mr_project_id):
    return rq.get(f'https://api.modrinth.com/v2/project/{mr_project_id}', params={'param': '1'},
                  headers=headers, verify=False).json()


# 处理最新上传的模组信息
def latest_upload():
    global LatestTime, headers, DuplicatesList, matchedSum, unmatchedSum, max_rowFIX, Json
    with Progress() as progress:
        task = progress.add_task("[red]正在初始化......", total=max_rowFIX)
        for NumID in range(2, max_rowFIX + 2):
            ModName = Vexl("C", NumID)
            # 更新任务进度
            NumProgress = "f[ {0}/{1} ]".format(NumID - 1, max_rowFIX)
            progress.update(task, advance=1,
                            description="[yellow]{0}正在处理：{1}".format(NumProgress, ModName))
            prog.setProgress(int(round(NumID * 100 / max_rowFIX, 0)))
            # 获取并处理Curseforge和Modrinth的数据
            try:
                CurseforgeID = Vexl("D", NumID)
                ModrinthID = Vexl("E", NumID)
                if CurseforgeID is not None:
                    Website = "Curseforge"
                    if not str(CurseforgeID).count("/"):
                        ID = CurseforgeID
                        Json = get_cfwidget_api_json(ID)['files']
                        fileIDs = [i["id"] for i in Json]
                        LatestTime = str(dict(Json[0])['uploaded_at'])
                    else:
                        for m in str(CurseforgeID).split("/"):
                            ID = m
                            Json = get_cfwidget_api_json(ID)['files']
                            fileIDs = [i["id"] for i in Json]
                            LatestTime += str(dict(Json[0])['uploaded_at']) + "  "

                elif ModrinthID is not None:
                    Website = "Modrinth"
                    if not str(ModrinthID).count("/"):
                        ID = ModrinthID
                        Json = get_modrinth_api_json(ID)
                        fileIDs = Json["versions"]
                        LatestTime = str(Json['updated'])
                    else:
                        for m in str(ModrinthID).split("/"):
                            ID = m
                            Json = get_modrinth_api_json(ID)
                            fileIDs += Json["versions"]
                            LatestTime += str(Json['updated']) + "  "
                else:
                    LatestTime = "读取失败"
            except Exception as Er:
                print(f"{NumID - 1} 出现错误:{Er}，跳过该项")
                LatestTime = "读取过程出现错误"
            finally:
                # 更新时间和处理重复项
                time_update("Latest", NumID, LatestTime)
                time_update("Json", NumID, fileIDs)
                if not check_duplicates(NumID):
                    DuplicatesList.append("{0}：{1} || {2} | {3}".format(NumID, ModName, LatestTime, Vexl("G", NumID)))
                    unmatchedSum += 1
                    pastJson = Vexl("K", NumID)
                    if pastJson:
                        fileID = list(set(fileIDs).difference(set(eval(pastJson))))
                        download_mod(Website, Vexl("B", NumID), processesTime, ID, fileID)
                else:
                    matchedSum += 1
                print(NumID - 1, ModName, LatestTime)
                LatestTime = fileIDs = ""

        progress.update(task, advance=1, description="[green][ √ ]已完成处理！")


# 初始化时更新时间
time_update("Latest", 1, processesTime)
time_update("Json", 1, processesTime)
# 开始处理最新上传的模组信息
latest_upload()

# 打印总结信息
print(f"匹配：{matchedSum}，不匹配{unmatchedSum}")
print(datetime.now(), "完成遍历")
if DuplicatesList:
    print(f"发现已更新的模组！：\n {DuplicatesList}")
else:
    print("没有发现已更新的模组")
# 保存Excel文件
try:
    wb.save(DatabasePath)
    print("所有改动已成功保存")
except Exception as E:
    print(f"保存出错: {E}")
# 更新任务栏状态
prog.setState('done')
