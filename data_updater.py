import config
from datetime import datetime
import PyTaskbar
import aiohttp
import asyncio
from toLog import toLog
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from rich.progress import Progress
from urllib3 import disable_warnings as urllib3_disable_warnings
import sys

if sys.platform == 'win32':
    asyncio.set_event_loop_policy(asyncio.WindowsSelectorEventLoopPolicy())
# Import custom downloader module
from mod_downloader import download_mod_metadata

config = config.Config()
DATABASE_PATH = config.DATABASE_PATH
download_enable = config.download_enable
blacklist_enabled = config.blacklist_enabled
POOL_SIZE = config.POOL_SIZE
CURSEFORGE_API_KEY = config.CURSEFORGE_API_KEY
if not CURSEFORGE_API_KEY:
    raise Exception('CURSEFORGE_API_KEY environment variable not set.')
# Set request headers
headers = config.headers
cf_headers = headers
cf_headers['x-api-key'] = CURSEFORGE_API_KEY

# Ignore urllib3 warnings
urllib3_disable_warnings()
# Define fill colors
green_fill = PatternFill(patternType='solid', fgColor='90EE90')  # Light green
red_fill = PatternFill(patternType='solid', fgColor='D94600')  # Red
blue_fill = PatternFill(patternType='solid', fgColor='A0C4FF')  # Light blue
no_fill = PatternFill(patternType='none')  # No fill

# Initialize global variables
max_rowFIX = unmatchedSum = matchedSum = 0
DuplicatesList = []
blacklist = config.blacklist if config.blacklist else []
processesTime = datetime.now()
validify_processesTime = str(processesTime).replace(" ", "+").replace(":", "-")

# Check if DATABASE_PATH environment variable is set
if not DATABASE_PATH:
    raise Exception('DATABASE_PATH environment variable not set.')

# Load Excel workbook
wb = load_workbook(DATABASE_PATH)
exl = wb.active

# Initialize taskbar progress bar
prog = PyTaskbar.Progress()
prog.init()
prog.setState('loading')


# Get Excel cell value by column and row
def Vexl(column, row):
    return exl[f"{column}{row}"].value


# Update time information
def time_update(type: str, row, time):
    Row_Str = str(row)
    if type == "Latest":
        exl["H" + Row_Str] = Vexl("G", row)
        exl["G" + Row_Str] = Vexl("F", row)
        exl["F" + Row_Str] = str(time)
    elif type == "Json":
        exl["L" + Row_Str] = Vexl("K", row)
        exl["K" + Row_Str] = Vexl("J", row)
        exl["J" + Row_Str] = str(time)


# Check for duplicates
def check_duplicates(num_id, latest_time):
    if latest_time != Vexl("F", num_id):
        exl[f"F{num_id}"].fill = red_fill
        return False
    else:
        exl[f"F{num_id}"].fill = green_fill
        return True


# Get max row count from Excel
for max_rowFIX, row in enumerate(exl, 1):
    if all(c.value is None for c in row):
        break
max_rowFIX -= 2


# Get Curseforge project API JSON data
async def get_curseforge_api_json_async(cf_project_id, session):
    retry_count = 0
    max_retries = config.TIMEOUT_RETRY
    while retry_count < max_retries:
        try:
            async with session.get(f'https://api.curseforge.com/v1/mods/{cf_project_id}/files?&pageSize=5000',
                                   headers=cf_headers, timeout=420) as response:
                return await response.json()
        except aiohttp.ClientResponseError as e:
            # Handle specific HTTP errors
            if e.status == 404:
                return {'data': []}
        except Exception as e:
            retry_count += 1
            if retry_count >= max_retries:
                toLog(f"Error getting Curseforge data for {cf_project_id} after {max_retries} attempts: {e}")
                return {'data': []}
            else:
                wait_time = 2 ** retry_count  # Exponential backoff
                toLog(f"Attempt {retry_count} failed for {cf_project_id}, retrying in {wait_time} seconds: {e}")
                await asyncio.sleep(wait_time)


# Get Modrinth project API JSON data
async def get_modrinth_api_json_async(mr_project_id, session):
    retry_count = 0
    max_retries = config.TIMEOUT_RETRY

    while retry_count < max_retries:
        try:
            async with session.get(f'https://api.modrinth.com/v2/project/{mr_project_id}/version',
                                   headers=headers, timeout=420) as response:

                return_json = await response.json()
                return {"versions": return_json}
        except aiohttp.ClientResponseError as e:
            if e.status == 404:
                return {"versions": []}
        except Exception as e:
            if "Expecting value: line 1 column 1 (char 0)" in str(e):
                return {"versions": []}
            retry_count += 1
            if retry_count >= max_retries:
                toLog(f"Error getting Modrinth data for {mr_project_id} after {max_retries} attempts: {e}")
                return {"versions": []}
            else:
                wait_time = 2 ** retry_count  # Exponential backoff
                toLog(
                    f"Attempt {retry_count} failed for {mr_project_id} because {e}, retrying in {wait_time} seconds...")
                await asyncio.sleep(wait_time)


async def add_unreachable_mod_to_blacklist(row: int):
    global blacklist
    # 如果F、G、H列均为“读取失败”，则列入黑名单
    if Vexl("J", row) == "[]" and Vexl("K", row) == "[]" and Vexl("L", row) == "[]":
        mcmod_id = Vexl("B", row)
        if mcmod_id not in config.blacklist:
            blacklist.append(mcmod_id)
            toLog(f"检测到多次读取百科ID为{mcmod_id}的模组失败，其将列入检测黑名单")
            config.write_config("Blacklist", blacklist)
            # 给涂上蓝色
            exl[f"B{row}"].fill = blue_fill


async def process_mod_async(num_id, session, progress, task):
    mod_name = Vexl("C", num_id)
    latest_time = website = used_id = None
    file_all = []

    try:
        curseforge_id = Vexl("D", num_id)
        modrinth_id = Vexl("E", num_id)
        # 给涂上白色
        exl[f"B{num_id}"].fill = no_fill

        if curseforge_id is not None:
            used_id = curseforge_id
            website = "Curseforge"
            if not str(used_id).count("/"):
                json_data = await get_curseforge_api_json_async(used_id, session)
                if 'data' in json_data and json_data['data']:
                    file_all = json_data['data']
                    latest_time = str(dict(file_all[0])['fileDate'])
                else:
                    if config.blacklist_enabled:
                        await add_unreachable_mod_to_blacklist(num_id)
                    raise ValueError("获取Curseforge数据为空")
            else:
                latest_time = ""
                for m in str(used_id).split("/"):
                    json_data = await get_curseforge_api_json_async(m, session)
                    if 'data' in json_data and json_data['data']:
                        file_all += json_data['data']
                        latest_time += str(dict(file_all[0])['fileDate']) + "  "
                    else:
                        if config.blacklist_enabled:
                            await add_unreachable_mod_to_blacklist(num_id)
                        raise ValueError("获取Curseforge数据为空")

        elif modrinth_id is not None:
            used_id = modrinth_id
            website = "Modrinth"
            if not str(used_id).count("/"):
                json_data = await get_modrinth_api_json_async(used_id, session)
                if 'versions' in json_data and json_data['versions']:
                    file_all = json_data["versions"]
                    latest_time = str(file_all[0]['date_published'])
                else:
                    if config.blacklist_enabled:
                        await add_unreachable_mod_to_blacklist(num_id)
                    raise ValueError("获取Modrinth数据为空")
            else:
                latest_time = ""
                for m in str(used_id).split("/"):
                    json_data = await get_modrinth_api_json_async(m, session)
                    if 'versions' in json_data and json_data['versions']:
                        file_all += json_data["versions"]
                        latest_time += str(file_all[0]['date_published']) + "  "
                    else:
                        if config.blacklist_enabled:
                            await add_unreachable_mod_to_blacklist(num_id)
                        raise ValueError("获取Modrinth数据为空")
        else:
            latest_time = "读取失败"
            if config.blacklist_enabled:
                await add_unreachable_mod_to_blacklist(num_id)

            website = None

        # check for duplicates
        matched = check_duplicates(num_id, latest_time)
        if not matched:
            toLog(f"{num_id - 1} {mod_name} 最新时间与上次不同，已标记为不匹配，其使用的id为：{used_id} ")
        return {
            "num_id": num_id,
            "mod_name": mod_name,
            "latest_time": latest_time,
            "file_all": file_all,
            "website": website if 'website' in locals() else None,
            "id": curseforge_id or modrinth_id,
            "matched": matched
        }

    except Exception as err:
        toLog(f"{num_id - 1} {mod_name} 出现错误:{err}，跳过该项")
        if config.blacklist_enabled:
            await add_unreachable_mod_to_blacklist(num_id)
        return {
            "num_id": num_id,
            "mod_name": mod_name,
            "latest_time": "读取过程出现错误",
            "file_all": "[]",
            "website": website if 'website' in locals() else None,
            "matched": False
        }
    finally:
        # Update progress regardless of success or failure
        progress.update(task, advance=1, description=f"[yellow]f[ {num_id - 1}/{max_rowFIX} ]处理中...")
        prog.setProgress(int(round(num_id * 100 / max_rowFIX, 0)))


# Modified download function to be async
async def save_mod_metadata_async(website, mcmod_id, time_str, signal_file_json):
    """
    根据提供的网站类型和项目ID下载Mod文件。
    :param signal_file_json: 该模组需要读取的单个文件元数据
    :param website: 网站名称（例如：Curseforge或Modrinth）
    :param mcmod_id: Mod的ID
    :param time_str: 保存目录的时间戳
    """
    loop = asyncio.get_event_loop()
    return await loop.run_in_executor(None, download_mod_metadata,
                                      website, mcmod_id, time_str, signal_file_json)


# Process latest uploaded mod information
async def latest_upload_async():
    global DuplicatesList, matchedSum, unmatchedSum
    matchedSum = unmatchedSum = 0

    with Progress() as progress:
        task = progress.add_task("[red]正在初始化......", total=max_rowFIX)

        # Create connection limit to avoid overwhelming servers
        connector = aiohttp.TCPConnector(limit=POOL_SIZE, ssl=True)
        async with aiohttp.ClientSession(connector=connector) as session:
            # Create a list of tasks
            tasks = []

            for num_id in range(2, max_rowFIX + 2):
                mod_mcmod_id = Vexl("B", num_id)
                if not blacklist_enabled or mod_mcmod_id not in blacklist:
                    task_obj = asyncio.create_task(process_mod_async(num_id, session, progress, task))
                    tasks.append(task_obj)

            # Execute all tasks and get results one by one to maintain order
            for task_obj in tasks:
                try:
                    result = await task_obj

                    # 最新时间存在差异时，执行以下内容
                    if not result["matched"] and result['latest_time'] != "读取过程出现错误":
                        DuplicatesList.append(
                            f"{result['num_id']}: {result['mod_name']} || {result['latest_time']} | {Vexl('F', result['num_id'])}")
                        unmatchedSum += 1

                        # Curseforge 和 Modrinth 都采用“ID”作为id键名
                        file_ids = [i["id"] for i in result["file_all"]]

                        if download_enable and 'website' in result and result['website']:
                            num_id = result['num_id']
                            toLog(f"开始处理 {result['mod_name']} 的最新上传信息")
                            mod_mcmod_id = Vexl("B", num_id)
                            # Look for most recent valid list as pastJson
                            pastIDData = None

                            try:
                                J_col = eval(Vexl("J", num_id))
                                K_col = eval(Vexl("K", num_id))
                                L_col = eval(Vexl("L", num_id))
                                if isinstance(J_col, list) and len(J_col) > 0:
                                    pastIDData = J_col
                                elif isinstance(K_col, list) and len(K_col) > 0:
                                    pastIDData = K_col
                                elif isinstance(L_col, list) and len(L_col) > 0:
                                    pastIDData = L_col
                            except Exception as e:
                                toLog(f"读取 {num_id} 行的历史文件ID数据失败，可能是格式错误：{e}")

                            if pastIDData:
                                try:
                                    different_fileIDs = set(file_ids) - set(pastIDData)
                                    for file_id in different_fileIDs:
                                        # 保存新文件的元数据
                                        # 找到“id” 为 file_id的项
                                        signal_file_json = next(
                                            (item for item in result["file_all"] if item["id"] == file_id), None)
                                        if signal_file_json:
                                            await save_mod_metadata_async(result['website'],
                                                                          mod_mcmod_id,
                                                                          validify_processesTime,
                                                                          signal_file_json)
                                except Exception as e:
                                    toLog(f"Error processing file differences: {e}")

                        # 更新改项的最新时间和file_ids
                        time_update("Latest", num_id, result['latest_time'])
                        time_update("Json", num_id, file_ids)
                    else:
                        matchedSum += 1

                except Exception as e:
                    toLog(f"Error processing result: {e}")

        progress.update(task, advance=0, description="[green][ √ ]已完成处理！")


# Main execution function
async def main():
    # Initialize time updates
    time_update("Latest", 1, processesTime)
    time_update("Json", 1, processesTime)

    # Report blacklist status
    if config.blacklist_enabled:
        if isinstance(config.blacklist, list):
            blacklisted_count = len(config.blacklist)
        else:
            config.blacklist = list()
            blacklisted_count = 0
        if blacklisted_count > 0:
            toLog(f"当前有 {blacklisted_count} 个模组在黑名单中，这些模组将被跳过处理")

    # Start processing latest uploaded mod information
    await latest_upload_async()

    # Print summary information
    toLog(f"匹配：{matchedSum}，不匹配：{unmatchedSum}")
    toLog("完成遍历")
    if DuplicatesList:
        toLog(f"发现已更新的模组：\n {DuplicatesList}")
        config.write_config("LastModified", str(validify_processesTime))
        config.write_config("Finished_upload", False)
    else:
        toLog("没有发现已更新的模组")

    # Save Excel file
    try:
        wb.save(DATABASE_PATH)
        toLog("所有改动已成功保存")

    except Exception as E:
        new_path = DATABASE_PATH.replace(".xlsx", f"_{validify_processesTime}.xlsx")
        toLog("保存出错: {0}\n 将保存至新文件{1}".format(E, new_path))
        wb.save(new_path)

    # Update taskbar status
    prog.setState('done')


# Run the async main function
if __name__ == "__main__":
    loop = asyncio.get_event_loop()
    try:
        loop.run_until_complete(main())
    finally:
        loop.run_until_complete(loop.shutdown_asyncgens())
        loop.close()
